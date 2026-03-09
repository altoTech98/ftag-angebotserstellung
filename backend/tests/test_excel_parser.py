"""Tests for excel_parser improvements – merged cells & multi-row headers."""

import pytest
import sys
import os
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

os.environ["TESTING"] = "true"
os.environ["ENVIRONMENT"] = "development"

import openpyxl


def _make_excel_with_merged_cells():
    """Create Excel with merged header cells (common in architect lists)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Türliste"

    # Row 1: merged group headers
    ws["A1"] = "Position"
    ws["B1"] = "Masse"
    ws.merge_cells("B1:C1")  # "Masse" spans B-C
    ws["D1"] = "Brandschutz"
    ws["E1"] = "Menge"

    # Row 2: sub-headers
    ws["A2"] = "Tür-Nr."
    ws["B2"] = "Breite [mm]"
    ws["C2"] = "Höhe [mm]"
    ws["D2"] = "Klasse"
    ws["E2"] = "Stk"

    # Data rows
    ws["A3"] = "T01"
    ws["B3"] = 900
    ws["C3"] = 2100
    ws["D3"] = "EI30"
    ws["E3"] = 2

    ws["A4"] = "T02"
    ws["B4"] = 1000
    ws["C4"] = 2200
    ws["D4"] = "EI60"
    ws["E4"] = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_excel_with_multi_row_header():
    """Excel where headers span 2 rows (group + detail)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Türmatrix"

    # Row 1: group headers
    ws["A1"] = "Tür"
    ws["B1"] = "Masse"
    ws["C1"] = ""  # part of Masse group
    ws["D1"] = "Schutz"
    ws["E1"] = ""  # part of Schutz group

    # Row 2: detail headers
    ws["A2"] = "Nr."
    ws["B2"] = "B [mm]"
    ws["C2"] = "H [mm]"
    ws["D2"] = "Brand"
    ws["E2"] = "RC"

    # Data
    ws["A3"] = "T01"
    ws["B3"] = 900
    ws["C3"] = 2100
    ws["D3"] = "EI30"
    ws["E3"] = "RC2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestUnmergeCells:
    """Tests for merged cell resolution."""

    def test_unmerge_fills_value(self):
        from services.excel_parser import unmerge_cells
        xlsx_bytes = _make_excel_with_merged_cells()
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        unmerge_cells(ws)
        # After unmerge, B1 and C1 should both have "Masse"
        assert ws["B1"].value == "Masse"
        assert ws["C1"].value == "Masse"

    def test_unmerge_preserves_non_merged(self):
        from services.excel_parser import unmerge_cells
        xlsx_bytes = _make_excel_with_merged_cells()
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        unmerge_cells(ws)
        assert ws["A1"].value == "Position"
        assert ws["D1"].value == "Brandschutz"


class TestMultiRowHeaders:
    """Tests for multi-row header detection and merging."""

    def test_combine_two_row_headers(self):
        from services.excel_parser import combine_multi_row_headers
        row1 = ["Position", "Masse", "Masse", "Brandschutz", "Menge"]
        row2 = ["Tür-Nr.", "Breite [mm]", "Höhe [mm]", "Klasse", "Stk"]
        combined = combine_multi_row_headers(row1, row2)
        assert "Breite" in combined[1] or "Masse" in combined[1]
        assert "Höhe" in combined[2] or "Masse" in combined[2]
        assert combined[0] in ("Tür-Nr.", "Position Tür-Nr.")

    def test_single_row_header_passthrough(self):
        from services.excel_parser import combine_multi_row_headers
        row1 = ["Tür-Nr.", "Breite [mm]", "Höhe [mm]", "Brandschutz"]
        combined = combine_multi_row_headers(row1, None)
        assert combined == row1


class TestMergedCellParsing:
    """Integration: parse Excel with merged cells end-to-end."""

    def test_parse_merged_cell_excel(self):
        from services.excel_parser import parse_tuerliste_bytes
        xlsx_bytes = _make_excel_with_merged_cells()
        result = parse_tuerliste_bytes(xlsx_bytes)
        doors = result["doors"]
        assert len(doors) >= 2
        # Check that dimensions were extracted
        found_dims = any(d.get("breite") == 900 for d in doors)
        assert found_dims, f"Expected breite=900, got: {[d.get('breite') for d in doors]}"

    def test_parse_multi_row_header_excel(self):
        from services.excel_parser import parse_tuerliste_bytes
        xlsx_bytes = _make_excel_with_multi_row_header()
        result = parse_tuerliste_bytes(xlsx_bytes)
        doors = result["doors"]
        assert len(doors) >= 1
        assert doors[0].get("brandschutz") is not None or doors[0].get("einbruchschutz") is not None
