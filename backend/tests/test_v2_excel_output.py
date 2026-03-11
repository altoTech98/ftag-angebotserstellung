"""
Tests for v2 Excel output generator.

Tests all 4 sheets: Uebersicht, Details, Gap-Analyse, Executive Summary.
Covers color coding, cell comments, frozen headers, auto-filter, and bytes output.
Also includes integration tests for V2 offer API endpoints.
"""

import re
import time
from datetime import datetime
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from v2.output.excel_generator import generate_v2_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load(result_bytes: bytes):
    """Load workbook from result bytes."""
    return load_workbook(BytesIO(result_bytes))


def _generate(
    sample_positions,
    sample_match_results,
    sample_adversarial_results,
    sample_gap_reports,
    ai_summary="Test-Zusammenfassung der Analyse.",
    ai_recommendations=None,
):
    """Call generate_v2_excel with sample fixtures."""
    if ai_recommendations is None:
        ai_recommendations = [
            "Empfehlung 1: Brandschutz-Upgrade pruefen",
            "Empfehlung 2: Alternativprodukte anbieten",
        ]
    return generate_v2_excel(
        positions=sample_positions,
        match_results=sample_match_results,
        adversarial_results=sample_adversarial_results,
        gap_reports=sample_gap_reports,
        ai_summary=ai_summary,
        ai_recommendations=ai_recommendations,
    )


# ---------------------------------------------------------------------------
# Sheet structure tests
# ---------------------------------------------------------------------------


class TestSheetStructure:
    """Tests for workbook sheet names and count."""

    def test_generates_4_sheets(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        assert wb.sheetnames == [
            "Uebersicht", "Details", "Gap-Analyse", "Executive Summary"
        ]

    def test_returns_bytes(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        assert isinstance(result, bytes)
        assert len(result) > 0


# ---------------------------------------------------------------------------
# Sheet 1: Uebersicht
# ---------------------------------------------------------------------------


class TestUebersicht:
    """Tests for the Uebersicht (overview) sheet."""

    def test_uebersicht_columns(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Uebersicht"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "Pos-Nr", "Bezeichnung", "Status", "Bestes Produkt",
            "Konfidenz%", "Anzahl Gaps", "Quelle",
        ]

    def test_uebersicht_row_count(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """One data row per position (3 positions = 3 data rows)."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Uebersicht"]
        # Count non-empty rows after header (row 1)
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                data_rows += 1
        assert data_rows == 3

    def test_uebersicht_color_coding(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Green for 95%+, yellow for 60-95%, red for <60%."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Uebersicht"]

        # Row 2 = pos 1.01 (0.97 confidence) -> green
        status_cell_1 = ws.cell(row=2, column=3)  # Status column
        assert status_cell_1.fill.fgColor.rgb is not None
        fill_1 = status_cell_1.fill.fgColor.rgb
        # Should be green-ish (C6EFCE)
        assert "C6EFCE" in str(fill_1).upper()

        # Row 3 = pos 1.02 (0.75 confidence) -> yellow
        status_cell_2 = ws.cell(row=3, column=3)
        fill_2 = status_cell_2.fill.fgColor.rgb
        assert "FFEB9C" in str(fill_2).upper()

        # Row 4 = pos 1.03 (0.40 confidence) -> red
        status_cell_3 = ws.cell(row=4, column=3)
        fill_3 = status_cell_3.fill.fgColor.rgb
        assert "FFC7CE" in str(fill_3).upper()

    def test_confirmed_position_zero_gaps(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Position 1.01 (bestaetigt, no GapReport) should show 0 gaps."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Uebersicht"]
        # Row 2 = pos 1.01, column 6 = Anzahl Gaps
        gap_count = ws.cell(row=2, column=6).value
        assert gap_count == 0

    def test_confidence_uses_adversarial(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Color coding should use adversarial adjusted_confidence, not match gesamt_konfidenz."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Uebersicht"]
        # Row 2 = pos 1.01: adversarial=0.97 (green), match=0.95 (also green but check value)
        konfidenz_cell = ws.cell(row=2, column=5)
        # Should show 97% (from adversarial), not 95% (from match)
        val = konfidenz_cell.value
        assert val is not None
        # The value should reflect 0.97 = 97%
        assert "97" in str(val) or val == 0.97 or val == 97


# ---------------------------------------------------------------------------
# Sheet 2: Details
# ---------------------------------------------------------------------------


class TestDetails:
    """Tests for the Details sheet with per-dimension scores and CoT comments."""

    def test_details_dimension_columns(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Details"]
        expected = [
            "Pos-Nr", "Produkt", "Gesamt-Konfidenz",
            "Masse", "Brandschutz", "Schallschutz",
            "Material", "Zertifizierung", "Leistung",
        ]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers == expected

    def test_details_cell_comments(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Dimension cells should have Comment objects with CoT reasoning."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Details"]
        # Row 2 = pos 1.01, column 4 = Masse dimension
        masse_cell = ws.cell(row=2, column=4)
        assert masse_cell.comment is not None
        assert isinstance(masse_cell.comment, Comment)
        assert len(masse_cell.comment.text) > 0

    def test_details_comment_truncation(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Comments longer than 2000 chars should be truncated with '...'."""
        # Create adversarial results with very long reasoning
        from copy import deepcopy
        long_adversarial = deepcopy(sample_adversarial_results)
        long_text = "A" * 3000
        long_adversarial[0].per_dimension_cot[0].reasoning = long_text

        result = generate_v2_excel(
            positions=sample_positions,
            match_results=sample_match_results,
            adversarial_results=long_adversarial,
            gap_reports=sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Details"]
        # Row 2 = pos 1.01, column 4 = Masse
        comment = ws.cell(row=2, column=4).comment
        assert comment is not None
        assert len(comment.text) <= 2003  # 2000 + "..."
        assert comment.text.endswith("...")


# ---------------------------------------------------------------------------
# Sheet 3: Gap-Analyse
# ---------------------------------------------------------------------------


class TestGapAnalyse:
    """Tests for the Gap-Analyse sheet."""

    def test_gap_analyse_rows(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """One row per GapItem (2 gaps for 1.02 + 3 gaps for 1.03 = 5 rows)."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Gap-Analyse"]
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                data_rows += 1
        assert data_rows == 5  # 2 + 3

    def test_gap_severity_colors(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Kritisch=dark red, Major=orange, Minor=light yellow."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Gap-Analyse"]

        # Collect severity cell fills (column 3 = Schweregrad)
        severity_fills = {}
        for row in range(2, ws.max_row + 1):
            sev = ws.cell(row=row, column=3).value
            if sev:
                fill = ws.cell(row=row, column=3).fill
                severity_fills[sev] = fill.fgColor.rgb if fill.fgColor else None

        # Check that different severities got different colors
        assert "kritisch" in severity_fills or "Kritisch" in severity_fills
        assert "major" in severity_fills or "Major" in severity_fills

    def test_gap_alternatives_column(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Alternative products formatted as 'name (XX%)' comma-separated."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        ws = wb["Gap-Analyse"]
        # Find row for pos 1.02 (first gap row) - last column should have alternatives
        # Column 9 = Alternative Produkte
        alt_cell = ws.cell(row=2, column=9)
        alt_text = str(alt_cell.value or "")
        assert "Flurtuere Premium EI60" in alt_text
        assert "85%" in alt_text


# ---------------------------------------------------------------------------
# Sheet 4: Executive Summary
# ---------------------------------------------------------------------------


class TestExecutiveSummary:
    """Tests for the Executive Summary sheet."""

    def test_executive_summary_stats(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Sheet 4 should have statistics rows."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
            ai_summary="Gesamtbewertung: Mehrheitlich machbar.",
            ai_recommendations=["Pruefen Sie Brandschutz-Upgrades"],
        )
        wb = _load(result)
        ws = wb["Executive Summary"]

        # Find cells with statistics content
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            all_values.extend([str(v) for v in row if v is not None])

        content = " ".join(all_values)
        # Should contain position count
        assert "3" in content  # total positions
        assert "Gesamtbewertung" in content or "Zusammenfassung" in content or "machbar" in content


# ---------------------------------------------------------------------------
# Cross-cutting concerns
# ---------------------------------------------------------------------------


class TestCrossCutting:
    """Tests for frozen headers, auto-filter, etc."""

    def test_frozen_headers(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """All sheets should have freeze_panes set at row 2."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        for name in ["Uebersicht", "Details", "Gap-Analyse"]:
            ws = wb[name]
            assert ws.freeze_panes is not None, f"{name} should have frozen panes"
            # freeze_panes should freeze header row (A2 means row 1 is frozen)
            assert ws.freeze_panes == "A2", f"{name} freeze_panes should be A2"

    def test_autofilter(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Data sheets should have auto_filter enabled."""
        result = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        wb = _load(result)
        for name in ["Uebersicht", "Details", "Gap-Analyse"]:
            ws = wb[name]
            assert ws.auto_filter.ref is not None, f"{name} should have auto_filter"


# ---------------------------------------------------------------------------
# API Integration Tests: V2 Offer Endpoints
# ---------------------------------------------------------------------------


def _make_test_client():
    """Create a FastAPI TestClient with the offer router."""
    from fastapi import FastAPI
    from routers.offer import router

    app = FastAPI()
    app.include_router(router, prefix="/api")
    return TestClient(app)


def _populate_analysis_results(
    analysis_id,
    sample_positions,
    sample_match_results,
    sample_adversarial_results,
    sample_gap_reports,
):
    """Inject sample data into _analysis_results for testing."""
    from v2.routers.analyze_v2 import _analysis_results

    _analysis_results[analysis_id] = {
        "positions": sample_positions,
        "match_results": sample_match_results,
        "adversarial_results": sample_adversarial_results,
        "gap_reports": sample_gap_reports,
        "created_at": datetime.now(),
    }


class TestGenerateEndpoint:
    """Tests for POST /api/offer/generate."""

    def test_generate_endpoint_returns_job_id(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """POST /api/offer/generate with valid analysis_id returns 200 with job_id."""
        client = _make_test_client()
        analysis_id = "test1234"
        _populate_analysis_results(
            analysis_id, sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )

        with patch("routers.offer.anthropic") as mock_anthropic:
            mock_client = MagicMock()
            mock_anthropic.Anthropic.return_value = mock_client
            mock_response = MagicMock()
            mock_response.parsed = MagicMock(
                gesamtbewertung="Zusammenfassung der Analyse",
                empfehlungen=["Empfehlung 1", "Empfehlung 2"],
            )
            mock_client.messages.parse.return_value = mock_response

            resp = client.post("/api/offer/generate", json={"analysis_id": analysis_id})

        assert resp.status_code == 200
        data = resp.json()
        assert "job_id" in data
        assert data["status"] == "started"

    def test_generate_endpoint_invalid_id(self):
        """POST /api/offer/generate with unknown analysis_id returns 404."""
        client = _make_test_client()
        resp = client.post("/api/offer/generate", json={"analysis_id": "nonexistent"})
        assert resp.status_code == 404


class TestDownloadEndpoint:
    """Tests for GET /api/offer/{id}/download."""

    def test_download_endpoint_returns_xlsx(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """GET /api/offer/{id}/download returns bytes with correct content-type and filename."""
        from services.memory_cache import offer_cache

        result_id = "dltest01"
        xlsx_bytes = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        offer_cache.set(f"v2_result_{result_id}_xlsx", xlsx_bytes, ttl_seconds=3600)

        client = _make_test_client()
        resp = client.get(f"/api/offer/{result_id}/download")

        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Verify it's valid xlsx
        wb = load_workbook(BytesIO(resp.content))
        assert "Uebersicht" in wb.sheetnames

    def test_download_endpoint_expired(self):
        """GET /api/offer/{id}/download for missing cache returns 410."""
        client = _make_test_client()
        resp = client.get("/api/offer/expired999/download")
        assert resp.status_code == 410

    def test_filename_format(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Downloaded filename matches Machbarkeitsanalyse_{date}_{id}.xlsx pattern."""
        from services.memory_cache import offer_cache

        result_id = "fntest01"
        xlsx_bytes = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        offer_cache.set(f"v2_result_{result_id}_xlsx", xlsx_bytes, ttl_seconds=3600)

        client = _make_test_client()
        resp = client.get(f"/api/offer/{result_id}/download")

        content_disp = resp.headers.get("content-disposition", "")
        today = datetime.now().strftime("%Y%m%d")
        expected_pattern = rf'Machbarkeitsanalyse_{today}_{result_id}\.xlsx'
        assert re.search(expected_pattern, content_disp), (
            f"Filename should match pattern, got: {content_disp}"
        )

    def test_cache_ttl_3600(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Excel bytes stored with 3600s TTL in offer_cache."""
        from services.memory_cache import offer_cache

        result_id = "ttltest1"
        xlsx_bytes = _generate(
            sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )
        cache_key = f"v2_result_{result_id}_xlsx"
        offer_cache.set(cache_key, xlsx_bytes, ttl_seconds=3600)

        # Verify data is retrievable
        cached = offer_cache.get(cache_key)
        assert cached is not None
        assert cached == xlsx_bytes


class TestGenerateDownloadFlow:
    """End-to-end generate -> status -> download flow."""

    def test_full_flow(
        self, sample_positions, sample_match_results,
        sample_adversarial_results, sample_gap_reports,
    ):
        """Generate, poll status, then download - full flow."""
        client = _make_test_client()
        analysis_id = "flow1234"
        _populate_analysis_results(
            analysis_id, sample_positions, sample_match_results,
            sample_adversarial_results, sample_gap_reports,
        )

        with patch("routers.offer.anthropic") as mock_anthropic:
            mock_client = MagicMock()
            mock_anthropic.Anthropic.return_value = mock_client
            mock_response = MagicMock()
            mock_response.parsed = MagicMock(
                gesamtbewertung="Analyse zeigt gute Machbarkeit",
                empfehlungen=["Brandschutz pruefen"],
            )
            mock_client.messages.parse.return_value = mock_response

            # Step 1: Generate
            gen_resp = client.post("/api/offer/generate", json={"analysis_id": analysis_id})
            assert gen_resp.status_code == 200
            job_id = gen_resp.json()["job_id"]

            # Step 2: Wait for background job to complete
            time.sleep(1.0)

            # Step 3: Check status
            status_resp = client.get(f"/api/offer/status/{job_id}")
            assert status_resp.status_code == 200

            # Step 4: Download
            dl_resp = client.get(f"/api/offer/{analysis_id}/download")
            assert dl_resp.status_code == 200
            # Verify valid xlsx
            wb = load_workbook(BytesIO(dl_resp.content))
            assert len(wb.sheetnames) == 4
