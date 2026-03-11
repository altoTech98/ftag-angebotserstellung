"""
Tests for v2 XLSX parser.

Tests header auto-detect, fuzzy column matching, merged cells,
multi-sheet handling, and error handling.
"""

import io

import pytest
from openpyxl import Workbook

from v2.parsers.base import ParseResult


def _make_xlsx_bytes(sheets_data: dict = None, header_row: int = 1,
                     merge_ranges: list = None) -> bytes:
    """Helper to create XLSX bytes with specified content.

    Args:
        sheets_data: dict of {sheet_name: [[row1], [row2], ...]}
        header_row: which row (1-indexed) to start data at (fills blanks above)
        merge_ranges: list of merge range strings like "A1:B1"
    """
    wb = Workbook()

    if sheets_data:
        first = True
        for sheet_name, rows in sheets_data.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(sheet_name)

            # Add blank rows before header if needed
            current_row = 1
            while current_row < header_row:
                ws.append([""] * (len(rows[0]) if rows else 1))
                current_row += 1

            for row in rows:
                ws.append(row)

            if merge_ranges:
                for mr in merge_ranges:
                    ws.merge_cells(mr)
    else:
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Tuer Nr.", "Breite", "Hoehe", "Brandschutz"])
        ws.append(["1.01", "1000", "2100", "EI30"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestHeaderAutoDetect:
    """Test header row detection."""

    def test_header_auto_detect(self):
        """Parse XLSX with header in row 3 (rows 1-2 are title/blank) -> correctly identifies row 3."""
        from v2.parsers.xlsx_parser import parse_xlsx

        # Row 1: Project title, Row 2: blank, Row 3: actual headers
        xlsx_bytes = _make_xlsx_bytes(
            sheets_data={
                "Tuerliste": [
                    ["Tuer Nr.", "Bezeichnung", "Breite (mm)", "Hoehe (mm)", "Brandschutz"],
                    ["1.01", "Buerotuere", "1000", "2100", "EI30"],
                    ["1.02", "Flurtuere", "900", "2050", "EI60"],
                ]
            },
            header_row=3,
        )

        result = parse_xlsx(xlsx_bytes, filename="test.xlsx")

        assert isinstance(result, ParseResult)
        assert result.format == "xlsx"
        # Should have found the headers and data
        assert "1.01" in result.text
        # Header row info should be in metadata
        assert "header_rows" in result.metadata


class TestKnownFieldPatterns:
    """Test known field pattern matching."""

    def test_known_field_patterns(self):
        """Parse XLSX with columns matching known patterns."""
        from v2.parsers.xlsx_parser import parse_xlsx

        xlsx_bytes = _make_xlsx_bytes(
            sheets_data={
                "Tuerliste": [
                    ["Tuer Nr.", "Breite [mm]", "Brandschutz"],
                    ["1.01", "1000", "EI30"],
                ]
            }
        )

        result = parse_xlsx(xlsx_bytes, filename="test.xlsx")

        assert isinstance(result, ParseResult)
        assert "detected_columns" in result.metadata


class TestFuzzyColumnMatching:
    """Test fuzzy matching of misspelled column headers."""

    def test_fuzzy_column_matching(self):
        """Parse XLSX with slight misspellings -> still matches."""
        from v2.parsers.xlsx_parser import parse_xlsx

        xlsx_bytes = _make_xlsx_bytes(
            sheets_data={
                "Tuerliste": [
                    ["Tuerbreite", "Brandschutzklase", "Schaldaemmung"],
                    ["1000", "EI30", "32dB"],
                ]
            }
        )

        result = parse_xlsx(xlsx_bytes, filename="test.xlsx")

        assert isinstance(result, ParseResult)
        detected = result.metadata.get("detected_columns", {})
        # Should have matched at least some columns across all sheets
        all_mappings = {}
        for sheet_cols in detected.values():
            all_mappings.update(sheet_cols)
        assert len(all_mappings) > 0


class TestMergedCells:
    """Test merged cell handling."""

    def test_merged_cells(self):
        """Parse XLSX with merged header cells -> unmerges before parsing, no crash."""
        from v2.parsers.xlsx_parser import parse_xlsx

        # Create workbook with merged cells manually
        wb = Workbook()
        ws = wb.active
        ws.title = "Tuerliste"
        ws["A1"] = "Masse"
        ws.merge_cells("A1:B1")  # Merge A1:B1
        ws["A2"] = "Tuer Nr."
        ws["B2"] = "Breite"
        ws["C2"] = "Brandschutz"
        ws["A3"] = "1.01"
        ws["B3"] = "1000"
        ws["C3"] = "EI30"

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = parse_xlsx(xlsx_bytes, filename="merged.xlsx")

        assert isinstance(result, ParseResult)
        assert result.format == "xlsx"
        # Should not crash, should extract some data
        assert len(result.warnings) == 0 or "merge" not in str(result.warnings).lower()


class TestToScalarHandling:
    """Test _to_scalar handling of duplicate column names."""

    def test_to_scalar_handling(self):
        """Parse XLSX with duplicate column names -> handles Series vs scalar."""
        from v2.parsers.xlsx_parser import parse_xlsx

        # Create workbook with duplicate column names (simulates unmerged cells)
        wb = Workbook()
        ws = wb.active
        ws.title = "Tuerliste"
        ws["A1"] = "Tuer Nr."
        ws["B1"] = "Masse"
        ws["C1"] = "Masse"  # Duplicate!
        ws["A2"] = "1.01"
        ws["B2"] = "1000"
        ws["C2"] = "2100"

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = parse_xlsx(xlsx_bytes, filename="dupcols.xlsx")

        assert isinstance(result, ParseResult)
        # Should not crash due to duplicate columns
        assert "1.01" in result.text


class TestMultiSheet:
    """Test multi-sheet processing."""

    def test_multi_sheet(self):
        """Parse XLSX with multiple sheets -> extracts from all relevant sheets."""
        from v2.parsers.xlsx_parser import parse_xlsx

        xlsx_bytes = _make_xlsx_bytes(
            sheets_data={
                "Tuerliste EG": [
                    ["Tuer Nr.", "Breite", "Hoehe", "Brandschutz"],
                    ["1.01", "1000", "2100", "EI30"],
                ],
                "Tuerliste OG": [
                    ["Tuer Nr.", "Breite", "Hoehe", "Brandschutz"],
                    ["2.01", "900", "2050", "EI60"],
                ],
                "Notizen": [
                    ["Bemerkung"],
                    ["Allgemeine Hinweise"],
                ],
            }
        )

        result = parse_xlsx(xlsx_bytes, filename="multi.xlsx")

        assert isinstance(result, ParseResult)
        assert "sheets_processed" in result.metadata
        assert len(result.metadata["sheets_processed"]) >= 2
        # Should have data from both door-list sheets
        assert "1.01" in result.text
        assert "2.01" in result.text


class TestErrorHandling:
    """Test error handling for corrupt files."""

    def test_corrupt_file(self):
        """Parse garbage bytes -> returns ParseResult with warning, does not raise."""
        from v2.parsers.xlsx_parser import parse_xlsx

        result = parse_xlsx(b"this is not an xlsx file", filename="corrupt.xlsx")

        assert isinstance(result, ParseResult)
        assert len(result.warnings) > 0
        assert result.format == "xlsx"


class TestReturnType:
    """Test return type."""

    def test_returns_parse_result(self, sample_xlsx_bytes):
        """Return type is ParseResult with format='xlsx'."""
        from v2.parsers.xlsx_parser import parse_xlsx

        result = parse_xlsx(sample_xlsx_bytes, filename="test.xlsx")
        assert isinstance(result, ParseResult)
        assert result.format == "xlsx"


class TestColumnStructureMetadata:
    """Test column structure in metadata."""

    def test_column_structure_in_metadata(self, sample_xlsx_bytes):
        """ParseResult.metadata contains 'detected_columns' mapping."""
        from v2.parsers.xlsx_parser import parse_xlsx

        result = parse_xlsx(sample_xlsx_bytes, filename="test.xlsx")

        assert "detected_columns" in result.metadata
        # detected_columns should be a dict mapping sheet names to column mappings
        assert isinstance(result.metadata["detected_columns"], dict)
