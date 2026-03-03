"""
Result Generator – Creates Tuerliste Excel im FTAG-Tuermatrix-Format + GAP Report.

Generates a 2-sheet Excel file:
  Sheet 1: "Tuermatrix-FTAG" – FTAG-format door matrix with product selections
  Sheet 2: "GAP-Report" – items that cannot be fulfilled

All functions return bytes (no disk writes).
"""

import io
import re
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def _auto_row_height(cell_values: list[str], col_widths: list[float], base_height: float = 15.0) -> float:
    """Estimate row height based on content and column widths."""
    max_lines = 1
    for val, width in zip(cell_values, col_widths):
        if not val:
            continue
        text = str(val)
        lines_in_text = text.count("\n") + 1
        chars_per_line = max(int(width * 1.2), 10)
        for segment in text.split("\n"):
            wrapped = max(1, -(-len(segment) // chars_per_line))
            lines_in_text += wrapped - 1
        max_lines = max(max_lines, lines_in_text)
    return max(base_height, min(max_lines * 15.0, 200.0))


def _clean_reason(entry: dict) -> str:
    """Build a human-readable reason text."""
    status = entry.get("status", "")
    gaps = entry.get("gap_items", [])
    products = entry.get("matched_products", [])

    if status == "unmatched" and not products:
        return "Kein passendes FTAG-Produkt gefunden"

    parts = []
    if gaps:
        for g in gaps:
            g = re.sub(r'^\[?\d+\]\s*\|?\s*', '', str(g))
            parts.append(g)
    elif status == "matched":
        parts.append("Anforderungen erfuellt")

    return "\n".join(parts) if parts else ""


def _extract_single_product_name(p: dict) -> str:
    """Extract a clean product name from a single product detail dict."""
    # Try direct column name from catalog detail
    for key in [
        "Türblatt / Verglasungsart / Rollkasten",
    ]:
        val = p.get(key, "")
        if val and val.strip():
            return str(val).strip()

    # Try via extended product info (catalog lookup by row_index)
    row_index = p.get("_row_index")
    if row_index is not None:
        try:
            from services.catalog_index import get_catalog_index
            catalog = get_catalog_index()
            ext = catalog.get_product_extended(row_index)
            tb = ext.get("tuerblatt", "")
            if tb:
                return tb
        except Exception:
            pass

    # Fallback: parse _compact text
    compact = p.get("_compact", "")
    if compact:
        # Extract product name from "[30] Category | ProductName | ..."
        parts = compact.split("|")
        if len(parts) >= 2:
            name = parts[1].strip()
            if name:
                return name

    return "-"


def _get_product_name(entry: dict) -> str:
    """Extract product name(s) (Türblatt) from all matched products.

    If multiple viable products exist, returns them joined with ' / '.
    E.g., 'Prestige 51 / Maxima 44 / Confort 59'
    """
    products = entry.get("matched_products", [])
    if not products:
        return "-"

    names = []
    for p in products:
        name = _extract_single_product_name(p)
        if name and name != "-" and name not in names:
            names.append(name)

    return " / ".join(names) if names else "-"


def _get_extended_product_info(entry: dict) -> dict:
    """Get extended product info from catalog (accessories, cost center, etc.)."""
    products = entry.get("matched_products", [])
    if not products:
        return {}
    p = products[0]
    row_index = p.get("_row_index")
    if row_index is None:
        compact = p.get("_compact", "")
        m = re.search(r'^\[(\d+)\]', compact)
        if m:
            row_index = int(m.group(1))
        else:
            return {}
    try:
        from services.catalog_index import get_catalog_index
        catalog = get_catalog_index()
        return catalog.get_product_extended(row_index)
    except Exception as e:
        logger.warning(f"Could not get extended product info: {e}")
        return {}


def _short_name(full_name: str) -> str:
    """Extract short product name from a full catalog column name.

    E.g. 'Band verdeckt Stumpf Tectus TE 540 (120 kg)' -> 'Tectus TE 540'
         'Dorma ITS 96 2-4' -> 'Dorma ITS 96 2-4'
         'Einsteckschloss Glutz' -> 'Glutz'
    """
    # Remove weight/spec in parentheses
    name = re.sub(r'\s*\(.*?\)\s*$', '', full_name).strip()
    # Remove common prefixes
    for prefix in [
        "Band sichtbar Stumpf ",
        "Band sichtbar Überschlag ",
        "Band verdeckt Stumpf ",
        "Band verdeckt Überschlag ",
        "Türschliessband verdeckt Stumpf ",
        "Zapfenband verdeckt Stumpf ",
        "Einsteckschloss ",
        "Mehrfachverriegelung ",
        "Falztreibriegel ",
        "Einlasskantenriegel ",
        "Einsteckfallenschloss ",
        "Einsteckriegelschloss ",
    ]:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return name.strip()


def _first_short(items: list, default: str = "-") -> str:
    """Get short name of first item in list, or default."""
    if not items:
        return default
    return _short_name(items[0])


def _normalize_dim(val) -> str:
    """Normalize a dimension value to mm string."""
    if not val:
        return ""
    try:
        n = float(val)
        if n <= 0:
            return ""
        if n < 20:
            return str(int(n * 1000))
        if n <= 400:
            return str(int(n * 10))
        return str(int(n))
    except (ValueError, TypeError):
        return str(val)


def _dim_to_ftag_cm(val, is_wall: bool = False) -> str:
    """Convert dimension to cm for FTAG output.

    For breite/hoehe: values > 400 are mm, divide by 10.
    For mauerdicke (is_wall=True): values > 50 are mm, divide by 10.
    """
    if not val:
        return ""
    try:
        n = float(val)
        if n <= 0:
            return ""
        threshold = 50 if is_wall else 400
        if n > threshold:
            n = n / 10  # mm -> cm
        return str(round(n))
    except (ValueError, TypeError):
        return str(val)


def _fmt_bs(val: str) -> str:
    """Format Brandschutz value for FTAG format."""
    if not val:
        return "-"
    val = str(val).strip()
    low = val.lower()
    if low in ("ohne", "keine", "-", "", "0", "nicht definiert"):
        return "-"
    # Standardize: EI30, EI60, etc.
    m = re.search(r'(ei|EI|t|T)\s*(\d+)', val)
    if m:
        return f"EI{m.group(2)}"
    return val


def generate_result_excel(
    matching: dict,
    requirements: dict,
    result_id: str,
) -> bytes:
    """Generate the Tuermatrix-FTAG Excel + GAP Report. Returns raw bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    red_fill = PatternFill("solid", fgColor="F8D7DA")
    alt_row_fill = PatternFill("solid", fgColor="F2F4F7")
    light_green = PatternFill("solid", fgColor="E8F5E9")
    meta_fill = PatternFill("solid", fgColor="E8EAF0")

    header_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F3A5F")
    subtitle_font = Font(name="Calibri", size=10, color="666666")
    normal_font = Font(name="Calibri", size=9)
    bold_font = Font(name="Calibri", size=10, bold=True)
    status_font_ok = Font(name="Calibri", size=9, bold=True, color="155724")
    status_font_warn = Font(name="Calibri", size=9, bold=True, color="856404")
    status_font_fail = Font(name="Calibri", size=9, bold=True, color="721C24")

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    center_v = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_center = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # ── Sheet 1: Tuermatrix-FTAG ────────────────────
    ws1 = wb.active
    projekt = requirements.get("projekt", "")
    metadata = requirements.get("metadata", {})
    summary = matching["summary"]
    total = summary["total_positions"]
    matched_cnt = summary["matched_count"]
    partial_cnt = summary["partial_count"]
    unmatch_cnt = summary["unmatched_count"]

    ws1.title = "Tuermatrix-FTAG"
    ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    # FTAG Türmatrix column config – exact match to FTAG Devi 49892 format
    # (header, width, alignment)
    columns = [
        ("Geschoss",                                                       9,   center_v),       # 1
        ("Tür-Nr. Architekt",                                             14,   left_center),    # 2
        ("Typenplan Architekt",                                           12,   center_v),       # 3
        ("Pos-Nr.- Sicherheitsplaner",                                    10,   left_center),    # 4
        ("Tür-Nr. Frank Türen AG",                                        14,   left_center),    # 5
        ("Betriebsauftrag Frank Türen AG",                                10,   left_center),    # 6
        ("Kostenträger Frank Türen AG",                                   22,   left_center),    # 7
        ("Zargentyp",                                                     12,   center_v),       # 8
        ("Mauertyp / LWB (Leichtbauzarge) / Mauerwerk (eingemörtelt)",    14,   center_v),       # 9
        ("Brandschutz, ohne BS / EI30 / EI60 / EI90",                    12,   center_v),       # 10
        ("Schallschutz, R`w+C in dB",                                     11,   center_v),       # 11
        ("Einbruchschutz, RC2-RC4",                                       11,   center_v),       # 12
        ("Klimaschutz",                                                    9,   center_v),       # 13
        ("Lichtmassbreite",                                               11,   center_v),       # 14
        ("Lichtmasshöhe",                                                 11,   center_v),       # 15
        ("Mauerdicke",                                                    10,   center_v),       # 16
        ("Türflügel, 1-flg / 2-flg",                                     10,   center_v),       # 17
        ("Türblatt",                                                      20,   left_center),    # 18
        ("Glaseinsatz",                                                   10,   center_v),       # 19
        ("Festverglasung EI30",                                            9,   center_v),       # 20
        ("Vollwand EI30",                                                  9,   center_v),       # 21
        ("Oberfläche Umfassung (Zarge / Rahmen)",                         16,   left_center),    # 22
        ("Oberfläche Türblatt",                                           16,   left_center),    # 23
        ("Bänder, 2 Stk/Element",                                        18,   left_center),    # 24
        ("Bandsicherung, 2 Stk / 3 Stk",                                 10,   center_v),       # 25
        ("Schlösser Gehflügel",                                           20,   left_center),    # 26
        ("Schlösser Standflügel",                                         14,   center_v),       # 27
        ("Zusatzfalle",                                                    9,   center_v),       # 28
        ("Kabelübergang steckbar 20-polig (Wanne-Stulp)",                  9,   center_v),       # 29
        ("Kabel zu Elektroschloss, 10m1 / 20m1",                          9,   center_v),       # 30
        ("Reedkontakt / Magnetkontakt, DMC 15 / DMC20",                   9,   center_v),       # 31
        ("Kabel zu Magnetkontakt, 10m1 / 20m1",                           9,   center_v),       # 32
        ("Elektrotüröffner, Eff-Eff 118 / Eff-Eff 143",                   9,   center_v),       # 33
        ("ElektroFluchtwerktüröffner, Eff-Eff 332 / Eff-Eff 331",         9,   center_v),       # 34
        ("Türschliesser",                                                 18,   left_center),    # 35
        ("Öffnungsbegrenzung / Rastfeststellung",                         10,   center_v),       # 36
        ("Haftmagnet",                                                     9,   center_v),       # 37
        ("Schiebetürautomat, Dorma / Gilgen",                              9,   center_v),       # 38
        ("Drehflügelautomat, Dorma ED 100 / Dorma ED 250",                9,   center_v),       # 39
        ("Radar / Sicherheitssensoren",                                    9,   center_v),       # 40
        ("Bodenabschluss, Senkdichtung Athmer SchallEX L-15 (52 dB)",    16,   left_center),    # 41
        ("Drückerpaar, Glutz / Mega",                                     10,   center_v),       # 42
        ("Drückerrosette, Glutz / Mega",                                   9,   center_v),       # 43
        ("Zylinderrosette, Glutz / Mega",                                  9,   center_v),       # 44
        ("Zylinder Bandseite, Zylinderachsmass bis Türflächs",             9,   center_v),       # 45
        ("Zylinder Bandgegenseite, Zylinderachsmass bis Türflächs",        9,   center_v),       # 46
        ("Wand-/ Bodenpuffer",                                             9,   center_v),       # 47
        ("Bemerkung Frank",                                               22,   left_top),       # 48
        ("Devi-Nr.",                                                      10,   left_center),    # 49
        ("Kalk-Nr.- Frank Türen AG",                                      10,   left_center),    # 50
        ("Total Kosten Frank Türen AG",                                   12,   right_center),   # 51
        ("Raum",                                                          20,   left_center),    # 52
        ("Status",                                                        14,   center_v),       # 53
        ("Hinweise",                                                      40,   left_top),       # 54
    ]

    col_widths_list = [c[1] for c in columns]

    for col_idx, (header, width, _) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = width

    # ── Title section (rows 1-3, no project metadata) ──
    last_col = get_column_letter(len(columns))
    ws1.merge_cells(f"A1:{last_col}1")
    ws1["A1"] = "LIS Türliste für Kalkulation mit Preisangaben, Frank Türen AG"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Row 2: Stand (date only – no project name, no bauherr/architekt)
    ws1["A2"] = "Stand"
    ws1["A2"].font = bold_font
    ws1.merge_cells(f"B2:{last_col}2")
    ws1["B2"] = datetime.now().strftime("%d.%m.%Y")
    ws1["B2"].font = subtitle_font
    ws1.row_dimensions[2].height = 18

    # Row 3: Summary
    ws1.merge_cells(f"A3:{last_col}3")
    ws1["A3"] = (
        f"Machbarkeitsanalyse: {total} Positionen  |  "
        f"{matched_cnt} machbar  |  {partial_cnt} teilweise  |  "
        f"{unmatch_cnt} nicht machbar  |  Match-Rate: {summary.get('match_rate', 0)}%"
    )
    ws1["A3"].font = Font(name="Calibri", size=9, bold=True, color="1F3A5F")
    ws1["A3"].fill = meta_fill
    ws1.row_dimensions[3].height = 20
    meta_row = 4

    # ── Header row ──
    HEADER_ROW = meta_row
    for col_idx, (header, _, _) in enumerate(columns, 1):
        cell = ws1.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_v
        cell.border = thin_border
    ws1.row_dimensions[HEADER_ROW].height = 40

    ws1.freeze_panes = f"A{HEADER_ROW + 1}"
    ws1.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{HEADER_ROW}"

    # ── Data rows ──────────────────────────────────
    all_positions = (
        matching.get("matched", [])
        + matching.get("partial", [])
        + matching.get("unmatched", [])
    )

    def _sort_key(entry):
        pos_str = entry.get("position", "")
        nums = re.findall(r"\d+", str(pos_str))
        return [int(n) for n in nums] if nums else [999999]

    all_positions.sort(key=_sort_key)

    row_num = HEADER_ROW + 1
    for i, entry in enumerate(all_positions):
        pos = entry.get("original_position", {})
        status = entry.get("status", "unmatched")

        # Status display
        if status == "matched":
            status_text = "MACHBAR"
            status_fill = green_fill
            s_font = status_font_ok
        elif status == "partial":
            status_text = "TEILWEISE"
            status_fill = yellow_fill
            s_font = status_font_warn
        else:
            status_text = "NICHT MACHBAR"
            status_fill = red_fill
            s_font = status_font_fail

        # Extended product info from catalog
        ext = _get_extended_product_info(entry)
        has_product = bool(ext)

        # Dimensions (normalize to cm for FTAG output)
        breite_cm = _dim_to_ftag_cm(pos.get("breite", ""))
        hoehe_cm = _dim_to_ftag_cm(pos.get("hoehe", ""))
        mauer_cm = _dim_to_ftag_cm(pos.get("mauerdicke", ""), is_wall=True)

        # Kostenträger from catalog
        kt = ext.get("kostentraeger", "")
        kat = str(entry.get("category", ""))
        if kt:
            kt_display = f"{kt}...{kat}" if kat else kt
        else:
            kt_display = "-"

        # Brandschutz
        bs_val = _fmt_bs(str(pos.get("brandschutz", "") or ""))

        # Schallschutz with dB suffix
        schall_raw = str(pos.get("schallschutz", "") or "").strip()
        if schall_raw and schall_raw not in ("-", "0", "ohne", "keine", "X"):
            schall_display = f"{schall_raw} dB" if "db" not in schall_raw.lower() else schall_raw
        elif schall_raw == "X":
            schall_display = "X"
        else:
            schall_display = "-"

        # Türflügel: "1" or "2" (short format like FTAG reference)
        fluegel_raw = ext.get("anzahl_fluegel", "")
        if fluegel_raw:
            if "2" in str(fluegel_raw):
                fluegel_display = "2"
            elif "1" in str(fluegel_raw):
                fluegel_display = "1"
            else:
                fluegel_display = str(fluegel_raw)
        else:
            fluegel_display = "-"

        # Hints / Gaps
        hints = _clean_reason(entry)

        # Zargentyp from catalog
        zargen = ext.get("zargen_types", [])
        zargen_display = zargen[0] if zargen else "-"

        # Mauertyp from position data
        mauertyp_raw = str(pos.get("mauertyp", "") or "").strip()
        mauertyp_display = mauertyp_raw if mauertyp_raw else "-"

        # Oberfläche from catalog (split: Umfassung vs Türblatt)
        oberfl_umf = ext.get("oberflaeche_umfassung", "") or "-"
        oberfl_tb = ext.get("oberflaeche_tuerblatt", "") or "-"

        # Verglasung
        verglasung = str(pos.get("verglasung", "") or pos.get("glas_typ", "") or "")
        glas_display = verglasung if verglasung else "-"

        # Accessories from catalog (real product names like FTAG reference)
        baender_display = _first_short(ext.get("baender", [])) if has_product else "-"
        schloss_display = _first_short(ext.get("schloesser", [])) if has_product else "-"
        boden_display = "X" if ext.get("bodenabschluss") else "-"

        # Türschliesser: only assign if door has Brandschutz requirement
        door_has_bs = bool(pos.get("brandschutz") and str(pos["brandschutz"]).strip()
                          not in ("", "-", "0", "ohne", "keine"))
        if has_product and door_has_bs and ext.get("tuerschliesser"):
            schliesser_display = _first_short(ext.get("tuerschliesser", []))
        else:
            schliesser_display = "-"

        # Build values matching FTAG Devi column order (54 columns)
        values = [
            str(pos.get("geschoss", "") or ""),                         # 1  Geschoss
            str(entry.get("position", "")),                             # 2  Tür-Nr. Architekt
            str(pos.get("tuertyp", "") or ""),                          # 3  Typenplan
            "",                                                          # 4  Pos-Nr. Sicherheitsplaner
            "",                                                          # 5  Tür-Nr. FTAG
            "",                                                          # 6  Betriebsauftrag FTAG
            kt_display,                                                  # 7  Kostenträger FTAG
            zargen_display,                                              # 8  Zargentyp
            mauertyp_display,                                            # 9  Mauertyp
            bs_val,                                                      # 10 Brandschutz
            schall_display,                                              # 11 Schallschutz
            str(pos.get("einbruchschutz", "") or "") or "-",            # 12 Einbruchschutz
            "-",                                                         # 13 Klimaschutz
            breite_cm or "-",                                            # 14 Lichtmassbreite
            hoehe_cm or "-",                                             # 15 Lichtmasshöhe
            mauer_cm or "-",                                             # 16 Mauerdicke
            fluegel_display,                                             # 17 Türflügel
            _get_product_name(entry),                                    # 18 Türblatt
            glas_display,                                                # 19 Glaseinsatz
            "-",                                                         # 20 Festverglasung EI30
            "-",                                                         # 21 Vollwand EI30
            oberfl_umf,                                                  # 22 Oberfläche Umfassung
            oberfl_tb,                                                   # 23 Oberfläche Türblatt
            baender_display,                                             # 24 Bänder
            "-",                                                         # 25 Bandsicherung
            schloss_display,                                             # 26 Schlösser Gehflügel
            "-",                                                         # 27 Schlösser Standflügel
            "-",                                                         # 28 Zusatzfalle
            "-",                                                         # 29 Kabelübergang
            "-",                                                         # 30 Kabel Elektroschloss
            "-",                                                         # 31 Reedkontakt
            "-",                                                         # 32 Kabel Magnetkontakt
            "-",                                                         # 33 Elektrotüröffner
            "-",                                                         # 34 ElektroFluchtwerktüröffner
            schliesser_display,                                          # 35 Türschliesser
            "-",                                                         # 36 Öffnungsbegrenzung
            "-",                                                         # 37 Haftmagnet
            "-",                                                         # 38 Schiebetürautomat
            "-",                                                         # 39 Drehflügelautomat
            "-",                                                         # 40 Radar
            boden_display,                                               # 41 Bodenabschluss
            "-",                                                         # 42 Drückerpaar
            "-",                                                         # 43 Drückerrosette
            "-",                                                         # 44 Zylinderrosette
            "-",                                                         # 45 Zylinder Bandseite
            "-",                                                         # 46 Zylinder Bandgegenseite
            "-",                                                         # 47 Wand-/Bodenpuffer
            hints if status != "matched" else "",                        # 48 Bemerkung Frank
            "",                                                          # 49 Devi-Nr.
            "",                                                          # 50 Kalk-Nr.
            "",                                                          # 51 Total Kosten
            str(pos.get("raum", pos.get("besonderheiten", "")) or ""),  # 52 Raum
            status_text,                                                 # 53 Status
            hints,                                                       # 54 Hinweise
        ]

        str_values = [str(v) if v else "" for v in values]
        row_height = _auto_row_height(str_values, col_widths_list)

        row_fill = alt_row_fill if i % 2 == 1 else None

        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_num, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = columns[col_idx - 1][2]
            if row_fill:
                cell.fill = row_fill

        # Status cell special formatting
        status_col_idx = next(
            (ci for ci, (h, _, _) in enumerate(columns, 1) if h == "Status"), 53
        )
        sc = ws1.cell(row=row_num, column=status_col_idx)
        sc.fill = status_fill
        sc.font = s_font
        sc.alignment = center_v

        ws1.row_dimensions[row_num].height = row_height
        row_num += 1

    # Summary row
    row_num += 1
    ws1.merge_cells(f"A{row_num}:{last_col}{row_num}")
    ws1.cell(row=row_num, column=1, value=(
        f"Zusammenfassung:  {matched_cnt} machbar  /  {partial_cnt} teilweise  /  "
        f"{unmatch_cnt} nicht machbar  (von {total} Positionen)  –  "
        f"Match-Rate: {summary.get('match_rate', 0)}%"
    )).font = bold_font
    ws1.cell(row=row_num, column=1).fill = meta_fill
    ws1.row_dimensions[row_num].height = 24

    ws1.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{row_num - 2}"

    # ── Sheet 2: GAP-Report ─────────────────────────
    ws2 = wb.create_sheet("GAP-Report")

    gap_columns = [
        ("Nr.",         5,   center_v),
        ("Tür-Nr.",    16,   left_center),
        ("Türtyp",     22,   left_center),
        ("Kategorie",  15,   left_center),
        ("B x H (mm)", 14,   center_v),
        ("Anforderung", 35,  left_top),
        ("Hinweis",    50,   left_top),
    ]
    gap_widths = [c[1] for c in gap_columns]

    for col_idx, (header, width, _) in enumerate(gap_columns, 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = width

    # Title
    last_gap_col = get_column_letter(len(gap_columns))
    ws2.merge_cells(f"A1:{last_gap_col}1")
    ws2["A1"] = "GAP-Report - Nicht erfuellbare Anforderungen"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells(f"A2:{last_gap_col}2")
    ws2["A2"] = f"Stand {datetime.now().strftime('%d.%m.%Y')}"
    ws2["A2"].font = subtitle_font
    ws2.row_dimensions[2].height = 20
    ws2.row_dimensions[3].height = 6

    # Header
    GAP_HEADER_ROW = 4
    for col_idx, (header, _, _) in enumerate(gap_columns, 1):
        cell = ws2.cell(row=GAP_HEADER_ROW, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_v
        cell.border = thin_border
    ws2.row_dimensions[GAP_HEADER_ROW].height = 26

    ws2.freeze_panes = "A5"

    gap_row = GAP_HEADER_ROW + 1
    gap_num = 0

    for entry in all_positions:
        gap_items = entry.get("gap_items", [])
        if not gap_items and entry.get("status") != "unmatched":
            continue

        pos = entry.get("original_position", {})

        # Dimensions
        breite = pos.get("breite", "")
        hoehe = pos.get("hoehe", "")
        dim_text = ""
        if breite and hoehe:
            try:
                b, h = float(breite), float(hoehe)
                if b < 20: b *= 1000
                elif b <= 400: b *= 10
                if h < 20: h *= 1000
                elif h <= 400: h *= 10
                dim_text = f"{int(b)} x {int(h)}"
            except (ValueError, TypeError):
                dim_text = f"{breite} x {hoehe}"

        tuertyp = str(pos.get("tuertyp", pos.get("beschreibung", ""))).strip()
        anforderung_parts = []
        if pos.get("brandschutz"):
            anforderung_parts.append(f"Brandschutz: {pos['brandschutz']}")
        if pos.get("schallschutz"):
            anforderung_parts.append(f"Schallschutz: {pos['schallschutz']}")
        if pos.get("einbruchschutz"):
            anforderung_parts.append(f"Einbruchschutz: {pos['einbruchschutz']}")
        anforderung = "\n".join(anforderung_parts)

        hinweis = "\n".join(str(g) for g in gap_items)

        gap_num += 1
        gap_values = [
            gap_num,
            str(entry.get("position", "")),
            tuertyp,
            str(entry.get("category", "")),
            dim_text,
            anforderung,
            hinweis,
        ]

        str_values = [str(v) if v else "" for v in gap_values]
        row_height = _auto_row_height(str_values, gap_widths)

        for col_idx, val in enumerate(gap_values, 1):
            cell = ws2.cell(row=gap_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = gap_columns[col_idx - 1][2]
            if gap_num % 2 == 0:
                cell.fill = alt_row_fill

        if entry.get("status") == "unmatched":
            ws2.cell(row=gap_row, column=1).fill = red_fill
        else:
            ws2.cell(row=gap_row, column=1).fill = yellow_fill

        ws2.row_dimensions[gap_row].height = row_height
        gap_row += 1

    if gap_num == 0:
        ws2.merge_cells(f"A5:{last_gap_col}5")
        ws2["A5"] = "Keine GAP-Positionen - alle Anforderungen erfuellbar!"
        ws2["A5"].font = Font(name="Calibri", size=12, bold=True, color="155724")
        ws2["A5"].fill = light_green
        ws2["A5"].alignment = center_v
        ws2.row_dimensions[5].height = 30

    # Summary
    gap_row += 1
    ws2.merge_cells(f"A{gap_row}:{last_gap_col}{gap_row}")
    ws2.cell(row=gap_row, column=1, value=(
        f"Total: {gap_num} GAP-Positionen von {total} Gesamtpositionen"
    )).font = bold_font
    ws2.cell(row=gap_row, column=1).fill = meta_fill
    ws2.row_dimensions[gap_row].height = 24

    ws2.auto_filter.ref = f"A{GAP_HEADER_ROW}:{last_gap_col}{gap_row - 2}"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
