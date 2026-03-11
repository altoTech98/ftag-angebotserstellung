"""
XLSX Parser - Extracts text from Excel files with header auto-detect and fuzzy column matching.

Ported from v1 excel_parser.py with clean interface returning ParseResult.
Handles merged cells, duplicate column names, multi-sheet workbooks.
Never raises exceptions to callers.
"""

import io
import logging
from difflib import SequenceMatcher

import pandas as pd

from v2.parsers.base import ParseResult

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Known field patterns for fuzzy column header matching (ported from v1)
# ---------------------------------------------------------------------------
KNOWN_FIELD_PATTERNS = {
    "tuer_nr": [
        "tür-nr", "tür nr", "türnummer", "tuer-nr", "tuer nr", "tuernr",
        "türnr", "pos", "position", "pos.", "pos-nr", "pos nr", "element-nr",
        "element nr", "elementnr", "door no", "door number",
    ],
    "geschoss": [
        "geschoss", "stockwerk", "etage", "og", "ug", "eg", "floor", "level",
        "ebene", "stock",
    ],
    "breite": [
        "fertiglichtmass breite", "breite", "b [mm]", "b[mm]", "b (mm)",
        "lichte breite", "tb", "türbreite", "tuerbreite", "breite mm",
        "breite [mm]", "lichte b", "width", "dm-b durchgang breite",
        "durchgang breite", "lw b", "lichtes mass b",
    ],
    "hoehe": [
        "fertiglichtmass höhe", "höhe", "hoehe", "h [mm]", "h[mm]", "h (mm)",
        "lichte höhe", "th", "türhöhe", "höhe mm", "höhe [mm]", "lichte h",
        "lichte höhe", "height", "dm-h durchgang höhe", "durchgang höhe",
        "lw h", "lichtes mass h",
    ],
    "rohbau_breite": [
        "rohbaumass breite", "rohbau breite", "rbm breite",
        "rm-b rohbaumass breite", "rbm b", "rohbaumass b",
        "rm-b rohbaumass- breite",
    ],
    "rohbau_hoehe": [
        "rohbaumass höhe", "rohbau höhe", "rbm höhe",
        "rm-h rohbaumass höhe", "rbm h", "rohbaumass h",
        "rm-h rohbaumass- höhe",
    ],
    "brandschutz": [
        "brandschutz", "feuerschutz", "bs", "feuerwiderstand",
        "brandschutzklasse", "brandschutzklase", "brand", "fire", "ei30",
        "ei60", "ei90", "feuerwiderstandsklasse", "brandschutzanforderung",
        "brand-/rauchschutz", "rauchschutz",
    ],
    "schallschutz": [
        "schallschutz", "ss", "rw", "schalldämmung", "schall",
        "schallschutzklasse", "schalldämmwert", "rw [db]", "rw[db]",
        "schalldämmwert (db)", "schaldaemmung",
    ],
    "einbruchschutz": [
        "einbruchschutz", "rc", "wk", "einbruch", "sicherheit",
        "widerstandsklasse", "einbruchsicherheit", "rc-klasse",
        "einbruch-widerstandsklasse",
    ],
    "tuertyp": [
        "türtyp", "tuertyp", "türart", "elementtyp",
        "tür typ", "door type", "konstruktion", "bauart", "türblatt",
    ],
    "beschlaege": [
        "beschläge", "beschlag", "drücker", "druecker", "schloss",
        "schliesser", "schliessung", "beschlaege", "bänder", "band",
        "türdrücker", "türschliesser", "türband",
    ],
    "oberflaechenbehandlung": [
        "oberfläche", "oberflaeche", "farbe", "ral", "beschichtung",
        "anstrich", "finish", "surface", "oberflächenbehandlung",
        "farbgebung", "ral-ton", "oberfl",
    ],
    "verglasung": [
        "verglasung", "verglast", "lichtausschnitt",
        "seitenteil", "glasausschnitt", "glasart", "durchsicht (glas",
    ],
    "menge": [
        "menge", "stk", "stück", "quantity", "qty",
    ],
    "besonderheiten": [
        "bemerkung", "besonderheit", "hinweis", "kommentar", "notiz",
        "anmerkung", "zusatz", "bemerkungen", "remarks", "notes",
    ],
    "raum": [
        "raum", "raumbezeichnung", "raumnummer", "raum-nr", "raum nr",
        "zimmer", "room", "raumname", "nutzung", "zugehöriger raum",
    ],
    "wandtyp": [
        "wandtyp", "wand", "wandkonstruktion", "wandaufbau", "mauerwerk",
        "wandstärke", "wanddicke", "mauertyp", "mauerart", "wandart",
    ],
    "schloss_typ": [
        "schloss", "schlosstyp", "schlossart", "verriegelung",
        "panikschloss", "einsteckschloss", "mehrfachverriegelung",
        "schliessung",
    ],
    "zylinder": [
        "zylinder", "zylindertyp", "profilzylinder", "schliesszylinder",
        "zylinderart",
    ],
    "glas_typ": [
        "glastyp", "glasart", "glasaufbau", "brandschutzglas",
        "lichtausschnitt typ", "glassorte",
    ],
    "schliessblech": [
        "schliessblech", "schliessstück", "schliessblech typ",
    ],
    "tuerschliesser": [
        "türschliesser", "schliesser", "obentürschliesser",
        "ots", "bodentürschliesser", "bts",
    ],
    "fluegel_anzahl": [
        "anzahl türflügel", "anzahl flügel", "flügelanzahl", "türflügel",
        "flügel", "1-flg", "2-flg",
        "einflügelig", "zweiflügelig", "anzahl fluegel", "anzahl tuerfluegel",
    ],
    "bandtyp": [
        "bandtyp", "bandart", "türband", "scharnier", "bandsicherung",
    ],
    "zargentyp": [
        "zargentyp", "zargenart", "zarge", "umfassungszarge",
        "blockzarge", "eckzarge", "umfassungsart", "umfassung", "zarge typ",
    ],
}

# Minimum number of known field matches to consider a sheet "door-list-like"
_MIN_DOOR_FIELDS = 3


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_scalar(value):
    """Ensure a value is a scalar, not a pandas Series.

    When a DataFrame has duplicate column names, row.get(col) may return
    a Series instead of a scalar. This takes the first element.
    """
    if isinstance(value, pd.Series):
        return value.iloc[0] if len(value) > 0 else None
    return value


def _fuzzy_ratio(a: str, b: str) -> float:
    """Simple fuzzy ratio using SequenceMatcher."""
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _best_field_match(column_name: str) -> tuple:
    """Find the best matching canonical field for a column header.

    Returns: (field_name, score) or (None, 0.0)
    """
    col_lower = column_name.lower().strip()
    if not col_lower or len(col_lower) < 2:
        return None, 0.0

    best_field = None
    best_score = 0.0

    for field, patterns in KNOWN_FIELD_PATTERNS.items():
        field_best = 0.0
        for pattern in patterns:
            # Exact match
            if pattern == col_lower:
                score = 0.98
            # Substring match (only for patterns >= 4 chars)
            elif len(pattern) >= 4 and pattern in col_lower:
                specificity = len(pattern) / len(col_lower)
                score = 0.80 + min(specificity * 0.15, 0.15)
            elif len(pattern) >= 4 and col_lower in pattern:
                score = 0.80
            else:
                # Fuzzy match
                if len(pattern) < 4:
                    continue
                ratio = _fuzzy_ratio(col_lower, pattern)
                score = ratio if ratio >= 0.60 else 0.0

            if score > field_best:
                field_best = score

        if field_best > best_score:
            best_score = field_best
            best_field = field

    return best_field, best_score


def match_columns(headers: list) -> dict:
    """Map each header to best matching known field using fuzzy matching.

    Args:
        headers: List of column header strings.

    Returns:
        dict of {header_text: canonical_field_name} for matched columns.
        Only includes matches with similarity > 0.65.
    """
    mapping = {}
    for header in headers:
        header_str = str(header).strip()
        if not header_str:
            continue
        field, score = _best_field_match(header_str)
        if field and score >= 0.65:
            mapping[header_str] = field
    return mapping


def unmerge_cells(ws) -> None:
    """Resolve merged cells in an openpyxl worksheet.

    Copies the merged cell's value into all cells of the range,
    then unmerges. Must be called BEFORE reading with pandas.
    """
    for merge_range in list(ws.merged_cells.ranges):
        min_row, min_col = merge_range.min_row, merge_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = value


def detect_header_row(ws, max_scan: int = 20) -> int:
    """Detect the header row in a worksheet by scoring rows.

    Scores based on: (a) number of non-empty cells, (b) matches against
    KNOWN_FIELD_PATTERNS, (c) ratio of text to numeric content.

    Args:
        ws: openpyxl worksheet.
        max_scan: Maximum rows to scan.

    Returns:
        1-indexed row number of the best header candidate.
    """
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                row_values.append(str(val).strip())
            else:
                row_values.append("")

        non_empty = [v for v in row_values if v]
        if not non_empty:
            continue

        # Score: number of non-empty cells
        count_score = len(non_empty)

        # Score: matches against known field patterns
        field_matches = 0
        for val in non_empty:
            field, score = _best_field_match(val)
            if field and score >= 0.65:
                field_matches += 1

        # Score: text ratio (headers are mostly text, not numbers)
        text_count = sum(1 for v in non_empty if v and not v.replace(".", "").replace(",", "").isdigit())
        text_ratio = text_count / len(non_empty) if non_empty else 0

        # Combined score: heavily weight field matches
        total_score = count_score * 0.3 + field_matches * 5.0 + text_ratio * 2.0

        if total_score > best_score:
            best_score = total_score
            best_row = row_idx

    return best_row


def _row_to_text(row, col_mapping: dict, headers: list) -> str:
    """Convert a DataFrame row to a text block with 'field: value' pairs.

    Uses canonical field names where available, original headers otherwise.
    """
    parts = []
    for header in headers:
        header_str = str(header).strip()
        if not header_str:
            continue
        value = _to_scalar(row.get(header))
        if value is not None and not pd.isna(value):
            val_str = str(value).strip()
            if val_str:
                canonical = col_mapping.get(header_str, header_str)
                parts.append(f"{canonical}: {val_str}")
    return " | ".join(parts)


def _df_to_markdown(df: pd.DataFrame) -> str:
    """Convert a DataFrame to a markdown table string."""
    if df.empty:
        return ""

    headers = [str(c).strip() for c in df.columns]
    rows = ["| " + " | ".join(headers) + " |"]
    rows.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for _, row in df.iterrows():
        cells = []
        for header in df.columns:
            val = _to_scalar(row.get(header))
            if val is not None and not pd.isna(val):
                cells.append(str(val).strip())
            else:
                cells.append("")
        rows.append("| " + " | ".join(cells) + " |")

    return "\n".join(rows)


# ---------------------------------------------------------------------------
# Main parse function
# ---------------------------------------------------------------------------

def parse_xlsx(content: bytes, filename: str = "") -> ParseResult:
    """Parse XLSX bytes into a ParseResult.

    Args:
        content: Raw XLSX file bytes.
        filename: Original filename for provenance tracking.

    Returns:
        ParseResult with extracted text and metadata. Never raises.
    """
    if not content:
        return ParseResult(
            text="",
            format="xlsx",
            page_count=0,
            warnings=["Empty file provided"],
            metadata={},
            source_file=filename,
        )

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        logger.warning(f"[XLSX] Could not open workbook {filename}: {e}")
        return ParseResult(
            text="",
            format="xlsx",
            page_count=0,
            warnings=[f"XLSX parsing failed: {str(e)}"],
            metadata={},
            source_file=filename,
        )

    text_parts = []
    tables_md = []
    detected_columns = {}
    header_rows = {}
    sheets_processed = []

    try:
        for ws in wb.worksheets:
            sheet_name = ws.title
            sheets_processed.append(sheet_name)

            # Step 1: Unmerge cells
            try:
                unmerge_cells(ws)
            except Exception as e:
                logger.debug(f"[XLSX] Unmerge failed for sheet {sheet_name}: {e}")

            # Step 2: Detect header row
            header_row = detect_header_row(ws)
            header_rows[sheet_name] = header_row

            # Step 3: Read into pandas DataFrame
            try:
                # Save workbook to buffer after unmerge, read specific sheet
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                df = pd.read_excel(
                    buf,
                    sheet_name=sheet_name,
                    header=header_row - 1,  # pandas uses 0-indexed
                    dtype=str,
                )
                df.columns = [str(c).strip() for c in df.columns]
                df = df.fillna("")

                # Drop completely empty rows
                df = df[df.apply(lambda row: any(str(v).strip() for v in row), axis=1)]

                if df.empty:
                    continue

            except Exception as e:
                logger.debug(f"[XLSX] Could not read sheet {sheet_name}: {e}")
                continue

            # Step 4: Match columns
            headers = list(df.columns)
            col_mapping = match_columns(headers)
            detected_columns[sheet_name] = col_mapping

            # Determine if this is a door-list sheet
            is_door_sheet = len(col_mapping) >= _MIN_DOOR_FIELDS

            # Step 5: Convert to text
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            if is_door_sheet:
                for _, row in df.iterrows():
                    row_text = _row_to_text(row, col_mapping, headers)
                    if row_text:
                        text_parts.append(row_text)
            else:
                # Basic text extraction for non-door sheets
                for _, row in df.iterrows():
                    cells = [str(_to_scalar(row.get(h))).strip()
                             for h in headers
                             if str(_to_scalar(row.get(h))).strip()
                             and str(_to_scalar(row.get(h))).strip() != "nan"]
                    if cells:
                        text_parts.append(" | ".join(cells))

            # Step 6: Create markdown table
            md = _df_to_markdown(df.head(50))  # Limit to 50 rows for tables list
            if md:
                tables_md.append(md)

    except Exception as e:
        logger.warning(f"[XLSX] Error processing workbook {filename}: {e}")
        return ParseResult(
            text="\n".join(text_parts) if text_parts else "",
            format="xlsx",
            page_count=len(sheets_processed),
            warnings=[f"Partial parsing: {str(e)}"],
            metadata={
                "detected_columns": detected_columns,
                "sheets_processed": sheets_processed,
                "header_rows": header_rows,
            },
            source_file=filename,
            tables=tables_md,
        )

    full_text = "\n".join(text_parts)

    metadata = {
        "detected_columns": detected_columns,
        "sheets_processed": sheets_processed,
        "header_rows": header_rows,
    }

    return ParseResult(
        text=full_text,
        format="xlsx",
        page_count=len(sheets_processed),
        warnings=[],
        metadata=metadata,
        source_file=filename,
        tables=tables_md,
    )
