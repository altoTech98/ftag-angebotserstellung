"""
V2 Excel Output Generator - 4-sheet professional Excel output.

Transforms v2 pipeline Pydantic schemas (MatchResult, AdversarialResult,
GapReport) into a polished, color-coded Excel file for the sales team.

Sheets:
  1. Uebersicht - One row per position, traffic-light status
  2. Details - Per-dimension scores with CoT cell comments
  3. Gap-Analyse - Individual gap items with severity colors
  4. Executive Summary - Statistics and AI assessment
"""

import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from v2.schemas.adversarial import AdversarialResult, DimensionCoT
from v2.schemas.gaps import GapReport, GapSeverity

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants: Color fills
# ---------------------------------------------------------------------------

# Traffic light fills for confidence status
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")

# Severity fills for gap items
KRITISCH_FILL = PatternFill("solid", fgColor="C00000")
KRITISCH_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
MAJOR_FILL = PatternFill("solid", fgColor="FFC000")
MINOR_FILL = PatternFill("solid", fgColor="FFF2CC")

# Header style
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=9)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")

# Alternating row shading
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

CENTER_V = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Dimension column order (must match MatchDimension/DimensionCoT names)
DIMENSION_ORDER = [
    "Masse", "Brandschutz", "Schallschutz",
    "Material", "Zertifizierung", "Leistung",
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _confidence_to_status(confidence: float) -> tuple[str, PatternFill, Font]:
    """Map confidence threshold to status text + fill + font."""
    if confidence >= 0.95:
        return (
            "Bestaetigt",
            GREEN_FILL,
            Font(name="Calibri", size=9, bold=True, color="155724"),
        )
    elif confidence >= 0.60:
        return (
            "Unsicher",
            YELLOW_FILL,
            Font(name="Calibri", size=9, bold=True, color="856404"),
        )
    else:
        return (
            "Abgelehnt",
            RED_FILL,
            Font(name="Calibri", size=9, bold=True, color="721C24"),
        )


def _severity_to_fill(severity: GapSeverity) -> tuple[PatternFill, Font]:
    """Map gap severity to color fill and font."""
    if severity == GapSeverity.KRITISCH:
        return KRITISCH_FILL, KRITISCH_FONT
    elif severity == GapSeverity.MAJOR:
        return MAJOR_FILL, NORMAL_FONT
    else:
        return MINOR_FILL, NORMAL_FONT


def _add_comment(cell, text: str, max_length: int = 2000) -> None:
    """Add a Comment to a cell, truncating at max_length."""
    if not text:
        return
    if len(text) > max_length:
        text = text[:max_length] + "..."
    cell.comment = Comment(text, "FTAG KI-Analyse", width=300, height=150)


def _style_header_row(ws, num_cols: int) -> None:
    """Apply header styling, freeze panes at A2, and set auto-filter."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_V
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}1"


def _auto_col_widths(ws) -> None:
    """Calculate and set column widths based on content (max 50, min 8)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        width = min(max(max_len + 2, 8), 50)
        ws.column_dimensions[col_letter].width = width


def _auto_row_height(
    cell_values: list[str], col_widths: list[float], base_height: float = 15.0
) -> float:
    """Estimate row height based on content and column widths."""
    max_lines = 1
    for val, width in zip(cell_values, col_widths):
        if not val:
            continue
        text = str(val)
        lines_in_text = text.count("\n") + 1
        chars_per_line = max(int(width * 1.2), 10)
        for segment in text.split("\n"):
            wrapped = max(1, -(-len(segment) // chars_per_line))
            lines_in_text += wrapped - 1
        max_lines = max(max_lines, lines_in_text)
    return max(base_height, min(max_lines * 15.0, 200.0))


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _write_uebersicht(wb, positions, match_lookup, adversarial_lookup, gap_lookup):
    """Write Sheet 1: Uebersicht (overview) with one row per position."""
    ws = wb.create_sheet("Uebersicht")
    ws.sheet_properties.tabColor = "00B050"  # green tab

    headers = [
        "Pos-Nr", "Bezeichnung", "Status", "Bestes Produkt",
        "Konfidenz%", "Anzahl Gaps", "Quelle",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for row_idx, pos in enumerate(positions, 2):
        pos_nr = pos.positions_nr

        # Get confidence from adversarial (preferred) or match
        adv = adversarial_lookup.get(pos_nr)
        match = match_lookup.get(pos_nr)
        if adv:
            confidence = adv.adjusted_confidence
        elif match and match.bester_match:
            confidence = match.bester_match.gesamt_konfidenz
        else:
            confidence = 0.0

        status_text, status_fill, status_font = _confidence_to_status(confidence)

        # Best product name
        product_name = "-"
        if adv and adv.bester_match:
            product_name = adv.bester_match.produkt_name
        elif match and match.bester_match:
            product_name = match.bester_match.produkt_name

        # Gap count
        gap_report = gap_lookup.get(pos_nr)
        gap_count = len(gap_report.gaps) if gap_report else 0

        # Source document
        quelle = "-"
        if pos.quellen:
            first_key = next(iter(pos.quellen))
            quelle = pos.quellen[first_key].dokument

        # Confidence display as percentage
        konfidenz_display = f"{confidence:.0%}"

        values = [pos_nr, pos.positions_bezeichnung or "-", status_text,
                  product_name, konfidenz_display, gap_count, quelle]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_CENTER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Status cell special formatting
        status_cell = ws.cell(row=row_idx, column=3)
        status_cell.fill = status_fill
        status_cell.font = status_font
        status_cell.alignment = CENTER_V

    _auto_col_widths(ws)


def _write_details(wb, positions, match_lookup, adversarial_lookup):
    """Write Sheet 2: Details with per-dimension scores and CoT comments."""
    ws = wb.create_sheet("Details")
    ws.sheet_properties.tabColor = "4472C4"  # blue tab

    headers = [
        "Pos-Nr", "Produkt", "Gesamt-Konfidenz",
    ] + DIMENSION_ORDER
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for row_idx, pos in enumerate(positions, 2):
        pos_nr = pos.positions_nr
        adv = adversarial_lookup.get(pos_nr)
        match = match_lookup.get(pos_nr)

        # Product name
        product_name = "-"
        if adv and adv.bester_match:
            product_name = adv.bester_match.produkt_name
        elif match and match.bester_match:
            product_name = match.bester_match.produkt_name

        # Gesamt-Konfidenz
        if adv:
            confidence = adv.adjusted_confidence
        elif match and match.bester_match:
            confidence = match.bester_match.gesamt_konfidenz
        else:
            confidence = 0.0

        ws.cell(row=row_idx, column=1, value=pos_nr).font = NORMAL_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=product_name).font = NORMAL_FONT
        ws.cell(row=row_idx, column=2).border = THIN_BORDER

        # Gesamt-Konfidenz cell with traffic light
        konf_cell = ws.cell(row=row_idx, column=3, value=f"{confidence:.0%}")
        _, konf_fill, konf_font = _confidence_to_status(confidence)
        konf_cell.fill = konf_fill
        konf_cell.font = konf_font
        konf_cell.border = THIN_BORDER
        konf_cell.alignment = CENTER_V

        # Build lookup dicts for dimension data
        # Short text from match DimensionScore.begruendung
        match_dim_lookup: dict[str, str] = {}
        if match and match.bester_match:
            for ds in match.bester_match.dimension_scores:
                match_dim_lookup[ds.dimension.value] = ds.begruendung

        # CoT reasoning from adversarial per_dimension_cot
        adv_dim_lookup: dict[str, DimensionCoT] = {}
        if adv and adv.per_dimension_cot:
            for cot in adv.per_dimension_cot:
                adv_dim_lookup[cot.dimension] = cot

        # Write dimension columns
        for dim_idx, dim_name in enumerate(DIMENSION_ORDER, 4):
            adv_cot = adv_dim_lookup.get(dim_name)
            match_text = match_dim_lookup.get(dim_name, "")

            if adv_cot:
                cell_text = f"{adv_cot.score:.0%} - {match_text}" if match_text else f"{adv_cot.score:.0%}"
                cot_text = adv_cot.reasoning
            elif match_text:
                # Find score from match
                score = 0.0
                if match and match.bester_match:
                    for ds in match.bester_match.dimension_scores:
                        if ds.dimension.value == dim_name:
                            score = ds.score
                            break
                cell_text = f"{score:.0%} - {match_text}"
                cot_text = match_text
            else:
                cell_text = "-"
                cot_text = ""

            dim_cell = ws.cell(row=row_idx, column=dim_idx, value=cell_text)
            dim_cell.font = NORMAL_FONT
            dim_cell.border = THIN_BORDER
            dim_cell.alignment = LEFT_CENTER

            if cot_text:
                _add_comment(dim_cell, cot_text)

        # Alternating row fill
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                if cell.fill == PatternFill():  # only if no special fill
                    cell.fill = ALT_ROW_FILL

    _auto_col_widths(ws)


def _write_gap_analyse(wb, gap_reports):
    """Write Sheet 3: Gap-Analyse with one row per GapItem."""
    ws = wb.create_sheet("Gap-Analyse")
    ws.sheet_properties.tabColor = "FF0000"  # red tab

    headers = [
        "Pos-Nr", "Dimension", "Schweregrad", "Anforderung",
        "Katalog", "Abweichung", "Kundenvorschlag",
        "Technischer Hinweis", "Alternative Produkte",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    row_idx = 2
    for report in gap_reports:
        # Build alternatives text for this position
        alt_parts = []
        for alt in report.alternativen:
            alt_parts.append(f"{alt.produkt_name} ({alt.teilweise_deckung:.0%})")
        alt_text = ", ".join(alt_parts) if alt_parts else "-"

        for gap in report.gaps:
            ws.cell(row=row_idx, column=1, value=report.positions_nr).font = NORMAL_FONT
            ws.cell(row=row_idx, column=1).border = THIN_BORDER

            ws.cell(row=row_idx, column=2, value=gap.dimension.value).font = NORMAL_FONT
            ws.cell(row=row_idx, column=2).border = THIN_BORDER

            # Severity with color
            sev_cell = ws.cell(row=row_idx, column=3, value=gap.schweregrad.value)
            sev_fill, sev_font = _severity_to_fill(gap.schweregrad)
            sev_cell.fill = sev_fill
            sev_cell.font = sev_font
            sev_cell.border = THIN_BORDER
            sev_cell.alignment = CENTER_V

            ws.cell(row=row_idx, column=4, value=gap.anforderung_wert).font = NORMAL_FONT
            ws.cell(row=row_idx, column=4).border = THIN_BORDER
            ws.cell(row=row_idx, column=5, value=gap.katalog_wert or "-").font = NORMAL_FONT
            ws.cell(row=row_idx, column=5).border = THIN_BORDER
            ws.cell(row=row_idx, column=6, value=gap.abweichung_beschreibung).font = NORMAL_FONT
            ws.cell(row=row_idx, column=6).border = THIN_BORDER
            ws.cell(row=row_idx, column=7, value=gap.kundenvorschlag or "-").font = NORMAL_FONT
            ws.cell(row=row_idx, column=7).border = THIN_BORDER
            ws.cell(row=row_idx, column=8, value=gap.technischer_hinweis or "-").font = NORMAL_FONT
            ws.cell(row=row_idx, column=8).border = THIN_BORDER
            ws.cell(row=row_idx, column=9, value=alt_text).font = NORMAL_FONT
            ws.cell(row=row_idx, column=9).border = THIN_BORDER

            row_idx += 1

    _auto_col_widths(ws)


def _write_executive_summary(
    wb, positions, match_lookup, adversarial_lookup,
    gap_reports, ai_summary_text: str, ai_recommendations: list[str],
):
    """Write Sheet 4: Executive Summary with statistics and AI assessment."""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = "FFC000"  # gold/amber tab

    # Count statistics
    total_pos = len(positions)
    confirmed = 0
    uncertain = 0
    rejected = 0
    for pos in positions:
        adv = adversarial_lookup.get(pos.positions_nr)
        if adv:
            if adv.adjusted_confidence >= 0.95:
                confirmed += 1
            elif adv.adjusted_confidence >= 0.60:
                uncertain += 1
            else:
                rejected += 1
        else:
            rejected += 1

    # Gap severity counts
    gap_kritisch = 0
    gap_major = 0
    gap_minor = 0
    gap_by_dim: dict[str, int] = {}
    for report in gap_reports:
        for gap in report.gaps:
            if gap.schweregrad == GapSeverity.KRITISCH:
                gap_kritisch += 1
            elif gap.schweregrad == GapSeverity.MAJOR:
                gap_major += 1
            else:
                gap_minor += 1
            dim_name = gap.dimension.value
            gap_by_dim[dim_name] = gap_by_dim.get(dim_name, 0) + 1

    row = 1
    # Title
    ws.merge_cells(f"A{row}:H{row}")
    title_cell = ws.cell(row=row, column=1, value="Executive Summary - KI-Angebotsanalyse")
    title_cell.font = TITLE_FONT
    ws.row_dimensions[row].height = 30

    row += 2  # blank row

    # Statistics section
    ws.cell(row=row, column=1, value="Statistik").font = BOLD_FONT
    row += 1

    stats = [
        ("Total Positionen", str(total_pos)),
        ("Matches (bestaetigt)", f"{confirmed} ({confirmed/total_pos*100:.0f}%)" if total_pos else "0"),
        ("Teilweise (unsicher)", f"{uncertain} ({uncertain/total_pos*100:.0f}%)" if total_pos else "0"),
        ("Kein Match (abgelehnt)", f"{rejected} ({rejected/total_pos*100:.0f}%)" if total_pos else "0"),
        ("", ""),
        ("Gaps nach Schweregrad", ""),
        ("  Kritisch", str(gap_kritisch)),
        ("  Major", str(gap_major)),
        ("  Minor", str(gap_minor)),
    ]

    if gap_by_dim:
        stats.append(("", ""))
        stats.append(("Gaps nach Dimension", ""))
        for dim, count in sorted(gap_by_dim.items()):
            stats.append((f"  {dim}", str(count)))

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1  # blank

    # AI Assessment
    ws.cell(row=row, column=1, value="KI-Zusammenfassung").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    summary_cell = ws.cell(row=row, column=1, value=ai_summary_text)
    summary_cell.font = NORMAL_FONT
    summary_cell.alignment = LEFT_TOP
    ws.row_dimensions[row].height = 60
    row += 2

    # AI Recommendations
    if ai_recommendations:
        ws.cell(row=row, column=1, value="Empfehlungen").font = BOLD_FONT
        row += 1
        for i, rec in enumerate(ai_recommendations, 1):
            ws.cell(row=row, column=1, value=f"{i}. {rec}").font = NORMAL_FONT
            row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    for col_letter in "CDEFGH":
        ws.column_dimensions[col_letter].width = 15


# ---------------------------------------------------------------------------
# Public function
# ---------------------------------------------------------------------------


def generate_v2_excel(
    positions: list,
    match_results: list,
    adversarial_results: list,
    gap_reports: list,
    ai_summary: str = "",
    ai_recommendations: list[str] | None = None,
) -> bytes:
    """Generate 4-sheet Excel from v2 pipeline results.

    Args:
        positions: ExtractedDoorPosition objects.
        match_results: MatchResult objects.
        adversarial_results: AdversarialResult objects.
        gap_reports: GapReport objects.
        ai_summary: Pre-generated AI summary text.
        ai_recommendations: Pre-generated recommendation list.

    Returns:
        Excel file as bytes (no disk writes).
    """
    if ai_recommendations is None:
        ai_recommendations = []

    # Build lookup dicts by positions_nr
    match_lookup = {mr.positions_nr: mr for mr in match_results}
    adversarial_lookup = {ar.positions_nr: ar for ar in adversarial_results}
    gap_lookup = {gr.positions_nr: gr for gr in gap_reports}

    wb = Workbook()

    # Write all 4 sheets
    _write_uebersicht(wb, positions, match_lookup, adversarial_lookup, gap_lookup)
    _write_details(wb, positions, match_lookup, adversarial_lookup)
    _write_gap_analyse(wb, gap_reports)
    _write_executive_summary(
        wb, positions, match_lookup, adversarial_lookup,
        gap_reports, ai_summary, ai_recommendations,
    )

    # Remove default "Sheet" created by Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Return bytes via BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
